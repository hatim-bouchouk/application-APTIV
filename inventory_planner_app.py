import tempfile
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import base64

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ─────────────────────────────────────────────────────────────────────────────
# 🎛️  CONFIGURATION CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_BOM_SHEET = "BOM"
DEFAULT_PLAN_SHEET = "plan"
DEFAULT_REQ_SHEET = "RM TOTAL REQUIREMENT"
DEFAULT_COV_SHEET = "coverage"
LOGO_PATH = "APTIV LG BG RM.png"  # place your logo file in the same folder (PNG/JPG)

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")

# ─────────────────────────────────────────────────────────────────────────────
# 🛠️  HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def detect_header_row(df_preview: pd.DataFrame, key_words: set[str] | None = None) -> int | None:
    key_words = key_words or {"Delphi PN", "Material"}
    for i, row in df_preview.iterrows():
        if key_words.intersection(set(row.fillna("").astype(str).str.strip())):
            return i
    return None


def is_excel_date(col) -> bool:
    if isinstance(col, (pd.Timestamp, datetime)):
        return True
    try:
        pd.to_datetime(col, errors="raise")
        return True
    except Exception:
        return False


def safe_float(x) -> float:
    try:
        if isinstance(x, str):
            x = x.replace(",", ".")
        return float(x)
    except Exception:
        return 0.0


def clean_comp(x: object) -> str:
    return str(x).strip().upper()

# ─────────────────────────────────────────────────────────────────────────────
# 📊  CORE PROCESSING PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_workbook(excel_path: Path,
                     bom_sheet: str,
                     plan_sheet: str,
                     req_sheet: str,
                     cov_sheet: str) -> tuple[pd.DataFrame, Path]:
    """Run the end‑to‑end raw‑material planning workflow and return the
    shortage DataFrame plus the path to the updated workbook."""

    # 1) ─── Build BOM dictionary ────────────────────────────────────────────
    bom_df = (
        pd.read_excel(excel_path, sheet_name=bom_sheet, dtype=object)
        .loc[:, ["Material", "Component", "Comp. Qty (BUn)"]]
        .dropna()
        .assign(**{"Comp. Qty (BUn)": lambda d: pd.to_numeric(d["Comp. Qty (BUn)"], errors="coerce")})
    )

    # 2) ─── Load plan sheet & detect header ─────────────────────────────────
    preview = pd.read_excel(excel_path, sheet_name=plan_sheet, header=None, nrows=10)
    hdr_row = detect_header_row(preview)
    if hdr_row is None:
        raise ValueError("Could not locate header row in plan sheet!")

    plan = pd.read_excel(excel_path, sheet_name=plan_sheet, header=hdr_row, dtype=object)
    date_cols = [c for c in plan.columns if is_excel_date(c)]
    if not date_cols:
        raise ValueError("No date columns detected in plan sheet.")

    # 3) ─── Compute total daily need (vectorised) ───────────────────────────
    plan_data = plan.iloc[hdr_row + 1:].copy()
    plan_data.rename(columns={plan.columns[1]: "FG"}, inplace=True)

    plan_long = (
        plan_data.melt(id_vars="FG", value_vars=date_cols, var_name="Date", value_name="QtyPlanned")
        .dropna(subset=["QtyPlanned"])
    )

    bom_small = (
        bom_df.rename(columns={"Material": "FG", "Comp. Qty (BUn)": "CompQty"})
        .loc[:, ["FG", "Component", "CompQty"]]
    )

    merged = (
        bom_small.merge(plan_long, on="FG", how="inner", copy=False)
        .assign(TotalNeed=lambda d: d["CompQty"] * d["QtyPlanned"])
    )

    tot_need_df = (
        merged.groupby(["Component", "Date"], as_index=False, sort=False)
        .agg(TotalNeed=("TotalNeed", "sum"))
    )

    # 4) ─── Write BOM explosion sheet ───────────────────────────────────────
    exploded_df = (
        tot_need_df.assign(Date=lambda d: pd.to_datetime(d["Date"]).dt.date)
        .loc[:, ["Component", "Date", "TotalNeed"]]
    )

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        exploded_df.to_excel(writer, sheet_name="BOM EXPLOSION", index=False)

    # 5) ─── Sync RM TOTAL REQUIREMENT sheet ────────────────────────────────
    wb = load_workbook(excel_path)
    ws_req = wb[req_sheet]

    # header row detection in RM sheet
    header_row = next((r for r in range(1, ws_req.max_row + 1)
                      if str(ws_req.cell(r, 1).value).strip().lower() == "component"), None)
    if header_row is None:
        raise ValueError("Cannot find 'Component' header in RM TOTAL REQUIREMENT sheet")

    ordered_dates = date_cols  # keep Excel order
    need_lookup = {(clean_comp(r.Component), pd.to_datetime(r.Date).date()): r.TotalNeed
                   for r in tot_need_df.itertuples(index=False)}

    date_cols_idx = list(range(2, 2 + len(ordered_dates)))

    for r in range(header_row + 1, ws_req.max_row + 1):
        comp = clean_comp(ws_req.cell(r, 1).value)
        if not comp:
            continue
        for i, col_idx in enumerate(date_cols_idx):
            day = pd.to_datetime(ordered_dates[i]).date()
            key = (comp, day)
            new_val = need_lookup.get(key)
            if new_val is None:
                continue
            cell = ws_req.cell(r, col_idx)
            old_val = safe_float(cell.value)
            if new_val > old_val:
                cell.value = new_val

    wb.save(excel_path)

    # 6) ─── Compute shortages from coverage sheet ───────────────────────────
    wb_cov = load_workbook(excel_path, data_only=True, read_only=True)
    ws_cov = wb_cov[cov_sheet]

    hdr = next(r for r in range(1, ws_cov.max_row + 1) if ws_cov.cell(r, 1).value == "APN")
    n_days = (ws_cov.max_column - 3) // 3
    idx_transit = [4 + n_days + i for i in range(n_days)]
    ordered_dates = [pd.to_datetime(d).date() for d in ordered_dates]

    fg_map = defaultdict(set)
    for r in merged.itertuples(index=False):
        fg_map[(clean_comp(r.Component), pd.to_datetime(r.Date).date())].add(str(r.FG))

    shortages = []
    for row in ws_cov.iter_rows(min_row=hdr + 1, max_row=ws_cov.max_row, max_col=3 + 3 * n_days, values_only=True):
        comp = clean_comp(row[0])
        if not comp:
            continue
        stock = safe_float(row[1])
        wip = safe_float(row[2])
        prev = stock + wip
        for i, day in enumerate(ordered_dates):
            if i:
                prev += safe_float(row[idx_transit[i - 1] - 1])
            need = need_lookup.get((comp, day), 0.0)
            balance = prev - need
            if balance < 0:
                fgs = ", ".join(sorted(fg_map[(comp, day)]) or ["–"])
                shortages.append((day, comp, fgs))
            prev = balance

    short_df = pd.DataFrame(shortages, columns=["Date", "Component", "FG Code"]).sort_values(["Date", "Component"])

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
        short_df.to_excel(xw, sheet_name="shortage", index=False)

    # 7) ─── Highlight shortages in plan sheet ───────────────────────────────
    wb_hl = load_workbook(excel_path)
    ws_plan = wb_hl[plan_sheet]

    header_row_excel = hdr_row + 1
    fg_col_idx = None
    date_to_col = {}

    for cell in ws_plan[header_row_excel]:
        val = cell.value
        if isinstance(val, str) and val.strip().lower() == "delphi pn":
            fg_col_idx = cell.col_idx
        try:
            dt = pd.to_datetime(val, errors="raise").date()
            date_to_col[dt] = cell.col_idx
        except Exception:
            pass

    if fg_col_idx is None:
        raise ValueError("Could not find 'Delphi PN' column in PLAN sheet header")

    short_fg_date = {(fg.strip(), pd.to_datetime(date).date())
                     for date, _comp, fgs in short_df.itertuples(index=False, name=None)
                     for fg in str(fgs).split(",")}

    for r in range(header_row_excel + 1, ws_plan.max_row + 1):
        fg_val = ws_plan.cell(r, fg_col_idx).value
        if fg_val is None:
            continue
        fg_clean = str(fg_val).strip()
        for d, col_idx in date_to_col.items():
            if (fg_clean, d) in short_fg_date:
                qty_cell = ws_plan.cell(r, col_idx)
                qty = safe_float(qty_cell.value)
                if qty > 0:
                    qty_cell.fill = RED

    wb_hl.save(excel_path)

    return short_df, excel_path

# ─────────────────────────────────────────────────────────────────────────────
# 🖥️  STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Material Bridge System", page_icon="🏭", layout="wide")

# ─── FIXED LOGO UPPER LEFT ──────────────────────────────────────────────────
if Path(LOGO_PATH).exists():
    with open(LOGO_PATH, "rb") as f:
        logo_data = base64.b64encode(f.read()).decode()
    st.markdown(
        f"""
        <style>
            .app-logo {{
                position: absolute;
                top: 0,1rem;
                left: 1rem;
                height: 130px;
            }}
        </style>
        <img src="data:image/png;base64,{logo_data}" class="app-logo">
        """,
        unsafe_allow_html=True
    )

# ─── HEADER TITLE ───────────────────────────────────────────────────────────
st.markdown("""
    <div style='margin-left: 10px; padding-top: 1rem;'>
        <h3 style='margin: 0;'></h3>
    </div>
    """, unsafe_allow_html=True)


with st.sidebar:
    
    st.title("Configuration")
    bom_sheet = st.text_input("BOM sheet name", value=DEFAULT_BOM_SHEET)
    plan_sheet = st.text_input("Plan sheet name", value=DEFAULT_PLAN_SHEET)
    req_sheet = st.text_input("Requirement sheet name", value=DEFAULT_REQ_SHEET)
    cov_sheet = st.text_input("Coverage sheet name", value=DEFAULT_COV_SHEET)
    run_btn = st.button("Run Analysis", type="primary", use_container_width=True)



uploaded_file = st.file_uploader("Upload your Excel workbook (.xlsx)", type=["xlsx"])

if run_btn and uploaded_file:
    with st.spinner("Processing workbook… this may take minutes ⏳"):
        # Save uploaded file to a temporary location
        tmp_dir = Path(tempfile.mkdtemp())
        saved_path = tmp_dir / uploaded_file.name
        saved_path.write_bytes(uploaded_file.getbuffer())

        try:
            short_df, updated_path = process_workbook(saved_path, bom_sheet, plan_sheet, req_sheet, cov_sheet)
        except Exception as e:
            st.error(f"❌ An error occurred: {e}")
            st.stop()

    st.success("✅ Analysis complete! Download your updated workbook below.")
    st.download_button("📥 Download updated workbook", data=updated_path.read_bytes(), file_name=f"processed_{uploaded_file.name}")

    st.subheader("🔎 Shortage Summary")
    st.dataframe(short_df, use_container_width=True)

else:
    st.info("Upload a workbook and press *Run Analysis* to start.")

# ─────────────────────────────────────────────────────────────────────────────
# 💅  CUSTOM CSS (optional – tweak as desired)
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
    <style>
    /* Rounded widgets & modern look */
    .stButton button {
        border-radius: 12px;
        font-weight: 600;
        padding: 0.5rem 1.25rem;
    }
    .stFileUploader, .stTextInput {
        border-radius: 10px !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
